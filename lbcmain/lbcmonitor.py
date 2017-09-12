from lbcapi import api
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path
import urlparse, json, sched, threading, time, pip, os

class Monitor:
    def __init__(self):
        path = os.getcwd()
        with open('./keys.json') as f:
            keys = json.load(f)
        self.__hmac_key = keys["READ"]["key"]
        self.__hmac_secret = keys["READ"]["secret"]
        self.__conn = api.hmac(self.__hmac_key, self.__hmac_secret)

    def get_public_ads(self, sellorbuy, currency, method):
        """Get public ads by connection

        Args:
            conn: Connection returned by api.hmac
            sellorbuy: string 'sell' or 'buy'
            currency: string of currency, example: 'CNY', 'USD'
            method: string of payment method, example: 'paypal', 'alipay'

        Returns:
            A dictionary responsed by server
        """
        json_response = self.__conn.call('GET', '/' + sellorbuy + '-bitcoins-online/' + currency + '/' + method + '/.json').json()
        try:
            return json_response['data']['ad_list']
        except:
            raise NoDataException(json_response['error']['message'])

    @staticmethod
    def get_first_ad(ads):
        """Get the first ad

        Args: 
            ads: Ads dictionary server returned

        Returns: 
            The first ad, always the best
        """
        return ads[1]['data']

    @staticmethod
    def get_max_value(ad1, ad2):
        """Get the ad with max price"""
        if float(ad1['temp_price']) > float(ad2['temp_price']):
            return ad1
        else:
            return ad2

    @staticmethod
    def get_min_value(ad1, ad2):
        """Get the ad with min price"""
        if float(ad1['temp_price']) < float(ad2['temp_price']):
            return ad1
        else:
            return ad2

    def usd_best_ad(self, sellorbuy):
        """Return best ad in USD"""
        return self.get_first_ad(self.get_public_ads(sellorbuy, 'USD', 'paypal'))

    def cny_best_ad(self, sellorbuy):
        """Return best ad in CNY"""
        cny_sell_ad_wechat = self.get_first_ad(self.get_public_ads(sellorbuy, 'CNY', 'wechat'))
        cny_sell_ad_alipay = self.get_first_ad(self.get_public_ads(sellorbuy, 'CNY', 'alipay'))
        if sellorbuy == 'sell':
            return self.get_max_value(cny_sell_ad_wechat, cny_sell_ad_alipay)
        else:
            return self.get_min_value(cny_sell_ad_wechat, cny_sell_ad_alipay)

    @staticmethod
    def add_to_xlsx(file_name, items):
        """Add a dictionary to the last row in xlsm

        Args:
            file_name: The file name to be created
            items: Include info of every columns
        """
        file_path = './' + file_name
        if os.path.isfile(file_path):
            wb = load_workbook(file_path)
        else:
            wb = Workbook()
            ws = wb.active
            ws['a1'] = 'Time'
            ws['b1'] = 'USD-sell price'
            ws['c1'] = ws['e1'] = ws['g1'] = 'amount'
            ws['d1'] = 'CNY-sell price'
            ws['f1'] = 'CNY-buy price'
        ws = wb.active
        row = ws.max_row + 1
        ws['a'+ str(row)] = time.ctime()
        ws['b'+ str(row)] = items['USD']['sell']['price']
        ws['c'+ str(row)] = items['USD']['sell']['max_amount']
        ws['d'+ str(row)] = items['CNY']['sell']['price']
        ws['e'+ str(row)] = items['CNY']['sell']['max_amount']
        ws['f'+ str(row)] = items['CNY']['buy']['price']
        ws['g'+ str(row)] = items['CNY']['buy']['max_amount']
        wb.save(file_path)
        wb.close()

    def check_value(self):
        """The main loop"""
        try:
            usd_sell_ad = self.usd_best_ad('sell')
            cny_sell_ad = self.cny_best_ad('sell')
            cny_buy_ad = self.cny_best_ad( 'buy')
            print "@time:", time.ctime()
            print "USD sell:", usd_sell_ad['temp_price'], "max value:", usd_sell_ad['max_amount_available']
            print "CNY sell:", cny_sell_ad['temp_price'], "max value:", cny_sell_ad['max_amount_available']
            print "CNY buy:", cny_buy_ad['temp_price'], "max value:", cny_buy_ad['max_amount_available']
            items = {'USD':
                     {'sell':
                      {'price':float(usd_sell_ad['temp_price']),'max_amount':int(usd_sell_ad['max_amount_available'])}},
                     'CNY':
                     {'sell':
                      {'price':float(cny_sell_ad['temp_price']),'max_amount':int(cny_sell_ad['max_amount_available'])},
                      'buy':
                      {'price':float(cny_buy_ad['temp_price']),'max_amount':int(cny_buy_ad['max_amount_available'])}}}
            self.add_to_xlsx("output.xlsx",items)
        except NoDataException as e:
            print "Exception:", e.message
        finally:
            threading.Timer(60.0, self.check_value).start()
        

    def start(self):
        """Start the main loop"""
        self.check_value()

class NoDataException(Exception):
    def __init__(self, message):
        self.message = message
    def __str__(self):
        return repr(self.message)