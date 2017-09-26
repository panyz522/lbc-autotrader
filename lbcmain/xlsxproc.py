import os, datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet import Worksheet
from config import Config

class XlsxProcessor(object):
    """XlsxProcesser maintain and update xlsm file to keep data"""
    
    def __init__(self):
        self.config = Config.get_xlsxconfig()
        self.file_path = os.path.join(self.config['path'],'doc',self.config['filename'])
        if os.path.isfile(self.file_path):
            self.wb = load_workbook(self.file_path)
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.row = 1 # current row
            self.crt_col = 1 # current colomn
            self.write( 'Time', 
                       'USD-sell price', 'Amount', 
                       'CNY-sell price', 'Amount', 
                       'CNY-buy price', 'Amount', 
                       'CNY-USD rate', 'CNY-USD netrate','Max amount')
        self.ws = self.wb.active
        self.row = self.ws.max_row + 1
        self.crt_col = 1

    def __enter__(self):
        return self

    def __exit__(self, type, val, tb):
        self.wb.save(self.file_path)
        self.wb.close()

    def add_to_xlsx(self, items):
        """Add a dictionary to the last row in xlsm

        Args:
            items: Include info of every columns
        """
        self.write("=DATEVALUE(AA" + str(self.row) + ")+TIMEVALUE(AA" + str(self.row) + ")")
        self.write(items['USD']['sell']['price'],
                   items['USD']['sell']['max_amount'],
                   items['CNY']['sell']['price'],
                   items['CNY']['sell']['max_amount'],
                   items['CNY']['buy']['price'],
                   items['CNY']['buy']['max_amount'],
                   items['rate']['C2U'],
                   items['net_rate']['C2U'],
                   items['available']['C2U']['inAd1'][1])
        self.crt_col = 27
        self.write(datetime.datetime.utcnow().strftime("%Y/%m/%d %H:%M:%S"))

    def write(self, *items):
        for item in items:
            self.ws.cell(row = self.row, column = self.crt_col, value = item)
            self.crt_col += 1